"""Report generation: PDF (reportlab), Excel (openpyxl), CSV."""
import io
import csv
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _fmt_dt(s: str) -> str:
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return s


def trips_to_csv(trips, classification_label: str) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow([
        "Date départ", "Date arrivée", "Conducteur", "Véhicule",
        "Départ", "Arrivée", "Distance (km)", "Durée (min)",
        "Carburant (L)", "Vitesse moyenne", "Vitesse max", "Type",
    ])
    for t in trips:
        w.writerow([
            _fmt_dt(t["start_time"]), _fmt_dt(t["end_time"]),
            t.get("driver_name", ""), t.get("vehicle_plate", ""),
            t.get("start_address", ""), t.get("end_address", ""),
            t.get("distance_km", 0), t.get("duration_min", 0),
            t.get("fuel_l", 0), t.get("avg_speed", 0), t.get("max_speed", 0),
            classification_label,
        ])
    return buf.getvalue().encode("utf-8-sig")


def trips_to_xlsx(trips, classification_label: str, title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trajets"

    headers = [
        "Date départ", "Date arrivée", "Conducteur", "Véhicule",
        "Départ", "Arrivée", "Distance (km)", "Durée (min)",
        "Carburant (L)", "Vit. moyenne", "Vit. max", "Type",
    ]
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="1F2937")
    ws.append([])

    ws.append(headers)
    header_row = ws.max_row
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col_idx)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2196F3")
        c.alignment = Alignment(horizontal="left", vertical="center")

    total_km = 0
    total_fuel = 0
    for t in trips:
        ws.append([
            _fmt_dt(t["start_time"]), _fmt_dt(t["end_time"]),
            t.get("driver_name", ""), t.get("vehicle_plate", ""),
            t.get("start_address", ""), t.get("end_address", ""),
            t.get("distance_km", 0), t.get("duration_min", 0),
            t.get("fuel_l", 0), t.get("avg_speed", 0), t.get("max_speed", 0),
            classification_label,
        ])
        total_km += t.get("distance_km", 0) or 0
        total_fuel += t.get("fuel_l", 0) or 0

    ws.append([])
    ws.append(["TOTAL", "", "", "", "", "", round(total_km, 1), "", round(total_fuel, 2), "", "", ""])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)

    from openpyxl.utils import get_column_letter
    widths = [18, 18, 22, 14, 40, 40, 12, 10, 12, 12, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def trips_to_pdf(trips, classification_label: str, title: str, subtitle: str = "") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1.0 * cm, rightMargin=1.0 * cm,
        topMargin=1.0 * cm, bottomMargin=1.2 * cm,
        title=title,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=16,
                                 alignment=0,  # left
                                 textColor=colors.HexColor("#0F172A"),
                                 spaceAfter=2)
    sub_style = ParagraphStyle("s", parent=styles["Normal"], fontSize=9,
                               textColor=colors.HexColor("#64748B"))
    cell_style = ParagraphStyle("c", parent=styles["Normal"], fontSize=8.2,
                                textColor=colors.HexColor("#0F172A"),
                                leading=10)
    addr_style = ParagraphStyle("a", parent=cell_style, fontSize=7.6,
                                textColor=colors.HexColor("#334155"),
                                leading=9.2)
    head_style = ParagraphStyle("h", parent=styles["Normal"], fontSize=8.5,
                                textColor=colors.white,
                                fontName="Helvetica-Bold")

    def P(text, style=cell_style):
        if text is None:
            text = ""
        return Paragraph(str(text).replace("&", "&amp;"), style)

    # Compute totals
    total_km = sum((t.get("distance_km") or 0) for t in trips)
    total_fuel = sum((t.get("fuel_l") or 0) for t in trips)
    total_min = sum((t.get("duration_min") or 0) for t in trips)
    hours = total_min // 60
    mins = total_min % 60

    # ----- Header block -----
    flow = []
    flow.append(Paragraph(title, title_style))
    meta_parts = []
    if subtitle:
        meta_parts.append(subtitle)
    meta_parts.append(f"{len(trips)} trajet{'s' if len(trips) > 1 else ''}")
    meta_parts.append(f"{total_km:,.1f} km".replace(",", "'"))
    meta_parts.append(f"{hours}h {mins:02d}min")
    meta_parts.append(f"{total_fuel:,.2f} L".replace(",", "'"))
    flow.append(Paragraph(" · ".join(meta_parts), sub_style))
    flow.append(Spacer(1, 0.35 * cm))

    # ----- Table -----
    header_row = [
        P("Date", head_style),
        P("Conducteur", head_style),
        P("Véhicule", head_style),
        P("Départ", head_style),
        P("Arrivée", head_style),
        P("Km", head_style),
        P("Durée", head_style),
        P("Carb. L", head_style),
        P("Type", head_style),
    ]

    data = [header_row]
    for t in trips:
        data.append([
            P(_fmt_dt(t.get("start_time", ""))),
            P(t.get("driver_name", "")),
            P(t.get("vehicle_plate", "")),
            P(t.get("start_address", ""), addr_style),
            P(t.get("end_address", ""), addr_style),
            P(f"{t.get('distance_km', 0):.1f}"),
            P(f"{t.get('duration_min', 0)} min"),
            P(f"{t.get('fuel_l', 0):.2f}"),
            P(classification_label),
        ])

    # Total row at the end
    bold_style = ParagraphStyle("b", parent=cell_style, fontName="Helvetica-Bold")
    data.append([
        P("", bold_style), P("", bold_style), P("", bold_style),
        P("", bold_style),
        P("TOTAL", bold_style),
        P(f"{total_km:.1f}", bold_style),
        P(f"{total_min} min", bold_style),
        P(f"{total_fuel:.2f}", bold_style),
        P("", bold_style),
    ])

    # Page width in landscape A4 minus margins ~ 27.7cm
    col_widths = [
        2.5 * cm,   # Date
        2.6 * cm,   # Conducteur
        2.0 * cm,   # Véhicule
        7.2 * cm,   # Départ
        7.2 * cm,   # Arrivée
        1.5 * cm,   # Km
        1.6 * cm,   # Durée
        1.5 * cm,   # Carb. L
        1.6 * cm,   # Type
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.6, colors.HexColor("#1565C0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E3F2FD")),
        ("LINEABOVE", (0, -1), (-1, -1), 0.8, colors.HexColor("#1976D2")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        # Right-align numeric columns
        ("ALIGN", (5, 1), (7, -1), "RIGHT"),
        # Center the Type column
        ("ALIGN", (8, 1), (8, -1), "CENTER"),
    ]))
    flow.append(table)

    flow.append(Spacer(1, 0.3 * cm))
    flow.append(Paragraph(
        f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} · "
        f"Logitrak — Livre de Bord · Données GPS officielles Navixy",
        ParagraphStyle("foot", parent=sub_style, fontSize=7.5,
                       textColor=colors.HexColor("#94A3B8")),
    ))

    doc.build(flow)
    return buf.getvalue()


def swiss_tax_report_pdf(stats: dict, year: int, owner: str = "") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("t", parent=styles["Title"], fontSize=20, textColor=colors.HexColor("#0F172A"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=13, textColor=colors.HexColor("#1976D2"))
    body = ParagraphStyle("b", parent=styles["BodyText"], fontSize=11, leading=16, textColor=colors.HexColor("#334155"))

    flow = []
    flow.append(Paragraph("Rapport Fiscal Annuel — Suisse", title_s))
    flow.append(Paragraph(f"Année fiscale : <b>{year}</b>", body))
    if owner:
        flow.append(Paragraph(f"Conducteur / Véhicule : <b>{owner}</b>", body))
    flow.append(Spacer(1, 0.5 * cm))
    flow.append(Paragraph("Synthèse kilométrique", h2))

    data = [
        ["Catégorie", "Valeur"],
        ["Kilomètres professionnels", f"{stats['pro_km']:,.1f} km".replace(",", "'")],
        ["Kilomètres personnels", f"{stats['perso_km']:,.1f} km".replace(",", "'")],
        ["Kilomètres totaux", f"{stats['total_km']:,.1f} km".replace(",", "'")],
        ["Pourcentage professionnel", f"{stats['pct_pro']:.1f} %"],
        ["Pourcentage personnel", f"{stats['pct_perso']:.1f} %"],
        ["Carburant professionnel (L)", f"{stats['pro_fuel']:.2f}"],
        ["Carburant personnel (L)", f"{stats['perso_fuel']:.2f}"],
    ]
    t = Table(data, colWidths=[9 * cm, 7 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    flow.append(t)
    flow.append(Spacer(1, 0.8 * cm))
    flow.append(Paragraph(
        "Ce document est généré automatiquement par Logitrak — Livre de Bord, "
        "à partir des données GPS officielles Navixy. Les kilomètres correspondent exactement "
        "aux distances Navixy. Conformité avec les exigences fiscales suisses (déduction privé/pro).",
        body,
    ))
    flow.append(Spacer(1, 0.3 * cm))
    flow.append(Paragraph(
        f"Document émis le {datetime.now().strftime('%d/%m/%Y à %H:%M')} — Logitrak SA, Genève.",
        ParagraphStyle("foot", parent=body, fontSize=9, textColor=colors.HexColor("#94A3B8")),
    ))

    doc.build(flow)
    return buf.getvalue()
