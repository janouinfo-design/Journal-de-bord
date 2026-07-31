"""Import de relevés carburant CSV / XLSX — parsing, mapping générique, normalisation.

Aucun format fournisseur n'est inventé : l'admin associe les colonnes de SON fichier
aux champs internes (mapping sauvegardable par fournisseur).
"""
import csv
import io
import re
from datetime import datetime, timezone

# Champs internes disponibles pour le mapping (label FR pour l'assistant)
INTERNAL_FIELDS = [
    {"key": "external_transaction_id", "label": "N° de transaction fournisseur"},
    {"key": "card_number", "label": "Numéro de carte (complet — jamais stocké)"},
    {"key": "card_last4", "label": "4 derniers chiffres de la carte"},
    {"key": "tx_datetime", "label": "Date et heure de la transaction"},
    {"key": "accounting_date", "label": "Date comptable"},
    {"key": "station_name", "label": "Station"},
    {"key": "station_address", "label": "Adresse de la station"},
    {"key": "country", "label": "Pays"},
    {"key": "station_lat", "label": "Latitude station"},
    {"key": "station_lng", "label": "Longitude station"},
    {"key": "product_type", "label": "Type de produit"},
    {"key": "quantity", "label": "Quantité"},
    {"key": "unit", "label": "Unité (L / kWh)"},
    {"key": "unit_price", "label": "Prix unitaire"},
    {"key": "amount_net", "label": "Montant HT"},
    {"key": "vat_amount", "label": "Montant TVA"},
    {"key": "vat_rate", "label": "Taux TVA (%)"},
    {"key": "amount_total", "label": "Montant TTC"},
    {"key": "currency", "label": "Devise"},
    {"key": "mileage", "label": "Kilométrage déclaré au terminal"},
    {"key": "vehicle_hint", "label": "Véhicule (plaque ou identifiant)"},
    {"key": "driver_hint", "label": "Chauffeur (nom ou identifiant)"},
    {"key": "invoice_ref", "label": "N° de facture / relevé"},
    {"key": "comment", "label": "Commentaire"},
]

# Auto-détection du mapping par mots-clés (insensible casse/accents)
_GUESS = {
    "external_transaction_id": ["transaction", "trx", "n° trans", "id"],
    "card_number": ["card number", "numero de carte", "numéro de carte", "kartennummer", "carte"],
    "card_last4": ["last4", "derniers chiffres"],
    "tx_datetime": ["date", "datum", "datetime", "date/heure"],
    "station_name": ["station", "site", "point de vente", "merchant"],
    "station_address": ["adresse", "address"],
    "country": ["pays", "country", "land"],
    "product_type": ["produit", "product", "article", "carburant", "fuel"],
    "quantity": ["quantit", "quantity", "menge", "litres", "liter", "volume"],
    "unit_price": ["prix unitaire", "unit price", "prix/l", "preis"],
    "amount_net": ["ht", "net", "netto"],
    "vat_amount": ["tva", "vat", "mwst"],
    "amount_total": ["ttc", "total", "brutto", "montant", "amount", "betrag"],
    "currency": ["devise", "currency", "währung", "waehrung"],
    "mileage": ["km", "kilom", "odometer", "mileage"],
    "vehicle_hint": ["vehicule", "véhicule", "vehicle", "plaque", "plate", "immatriculation", "kennzeichen"],
    "driver_hint": ["chauffeur", "driver", "conducteur", "fahrer"],
    "invoice_ref": ["facture", "invoice", "relev", "rechnung"],
}

_DT_FORMATS = [
    "%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%d.%m.%Y",
    "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
    "%d.%m.%y %H:%M", "%d.%m.%y",
]

_PRODUCT_MAP = {
    "diesel": "diesel", "gazole": "diesel", "gasoil": "diesel", "b7": "diesel",
    "essence": "essence", "sans plomb": "essence", "unleaded": "essence", "e10": "essence",
    "95": "essence", "98": "essence", "benzin": "essence", "super": "essence",
    "adblue": "adblue",
    "electric": "electric", "recharge": "electric", "charge": "electric", "kwh": "electric",
    "strom": "electric", "élec": "electric", "elec": "electric",
}


def parse_file(data: bytes, filename: str):
    """Retourne (columns, rows) — rows = list[dict col -> str]."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xls")):
        return _parse_xlsx(data)
    return _parse_csv(data)


def _parse_csv(data: bytes):
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Encodage du fichier non reconnu")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        class _D(csv.excel):
            delimiter = ";"
        dialect = _D
    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not all_rows:
        raise ValueError("Fichier vide")
    columns = [c.strip() for c in all_rows[0]]
    rows = []
    for r in all_rows[1:]:
        rows.append({columns[i]: (r[i].strip() if i < len(r) else "") for i in range(len(columns))})
    return columns, rows


def _parse_xlsx(data: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        raise ValueError("Fichier vide")
    columns = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(header)]
    rows = []
    for r in it:
        if r is None or all(v is None or str(v).strip() == "" for v in r):
            continue
        d = {}
        for i, col in enumerate(columns):
            v = r[i] if i < len(r) else None
            if isinstance(v, datetime):
                d[col] = v.strftime("%Y-%m-%d %H:%M:%S")
            else:
                d[col] = str(v).strip() if v is not None else ""
        rows.append(d)
    wb.close()
    return columns, rows


def guess_mapping(columns: list[str]) -> dict:
    """Propose automatiquement un mapping colonne -> champ interne."""
    mapping = {}
    used = set()
    for field, keywords in _GUESS.items():
        for col in columns:
            if col in used:
                continue
            low = col.lower()
            if any(k in low for k in keywords):
                mapping[col] = field
                used.add(col)
                break
    return mapping


def _num(val):
    if val is None or val == "":
        return None
    s = str(val).replace("'", "").replace("\u00a0", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "") if s.rindex(".") > s.rindex(",") else s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except ValueError:
        return None


def _dt(val):
    if not val:
        return None
    s = str(val).strip()
    try:
        d = datetime.fromisoformat(s)
        return (d if d.tzinfo else d.replace(tzinfo=timezone.utc)).isoformat()
    except ValueError:
        pass
    for fmt in _DT_FORMATS:
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            continue
    return None


def normalize_product(raw: str) -> str:
    low = (raw or "").strip().lower()
    for k, v in _PRODUCT_MAP.items():
        if k in low:
            return v
    return "other" if low else ""


def normalize_row(raw: dict, mapping: dict) -> tuple[dict, list[str]]:
    """mapping = {colonne_fichier: champ_interne}. Retourne (tx_partiel, erreurs)."""
    tx = {}
    errors = []
    for col, field in mapping.items():
        if not field or field == "ignore":
            continue
        tx[field] = raw.get(col, "")

    out = {}
    out["external_transaction_id"] = (str(tx.get("external_transaction_id") or "").strip() or None)
    out["card_number"] = str(tx.get("card_number") or "").strip() or None
    last4_src = str(tx.get("card_last4") or "").strip() or (out["card_number"] or "")
    digits = re.sub(r"\D", "", last4_src)
    out["card_last4"] = digits[-4:] if len(digits) >= 4 else None

    out["tx_datetime"] = _dt(tx.get("tx_datetime"))
    if not out["tx_datetime"]:
        errors.append("Date/heure manquante ou invalide")
    out["accounting_date"] = _dt(tx.get("accounting_date"))

    out["station_name"] = str(tx.get("station_name") or "").strip() or None
    out["station_address"] = str(tx.get("station_address") or "").strip() or None
    out["country"] = (str(tx.get("country") or "").strip().upper()[:2] or None)
    out["station_lat"] = _num(tx.get("station_lat"))
    out["station_lng"] = _num(tx.get("station_lng"))

    out["product_type"] = normalize_product(str(tx.get("product_type") or "")) or None
    out["quantity"] = _num(tx.get("quantity"))
    unit_raw = str(tx.get("unit") or "").strip().lower()
    out["unit"] = ("kWh" if "kwh" in unit_raw or out["product_type"] == "electric"
                   else "L" if not unit_raw or "l" in unit_raw else "unit")
    out["unit_price"] = _num(tx.get("unit_price"))
    out["amount_net"] = _num(tx.get("amount_net"))
    out["vat_amount"] = _num(tx.get("vat_amount"))
    out["vat_rate"] = _num(tx.get("vat_rate"))
    out["amount_total"] = _num(tx.get("amount_total"))
    if out["amount_total"] is None:
        errors.append("Montant TTC manquant ou invalide")
    out["currency"] = (str(tx.get("currency") or "").strip().upper() or "CHF")
    out["mileage"] = _num(tx.get("mileage"))
    out["vehicle_hint"] = str(tx.get("vehicle_hint") or "").strip() or None
    out["driver_hint"] = str(tx.get("driver_hint") or "").strip() or None
    out["invoice_ref"] = str(tx.get("invoice_ref") or "").strip() or None
    out["comment"] = str(tx.get("comment") or "").strip() or None

    # cohérence quantité × prix unitaire vs total (tolérance 5 %)
    if (out["quantity"] and out["unit_price"] and out["amount_total"]
            and out["amount_total"] > 0):
        computed = out["quantity"] * out["unit_price"]
        if abs(computed - out["amount_total"]) / out["amount_total"] > 0.05:
            errors.append(f"Montant incohérent (qté × prix = {computed:.2f} ≠ TTC {out['amount_total']:.2f})")

    return out, errors
