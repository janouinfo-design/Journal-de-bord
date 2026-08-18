"""Phase 6 — Fines export module.

Generates PDF / Excel / CSV files for the Fines management module. Reuses the
same filter parameters as `GET /api/livre/fines` so a user can export exactly
what they see in the table.

Why separate from the existing reports module: this is purely the fines
domain and shipping it inline keeps the module self-contained.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

# Status / type French labels — mirror of frontend constants
STATUS_LABEL = {
    "received": "Reçue", "to_analyze": "À analyser",
    "driver_to_identify": "Conducteur à identifier",
    "awaiting_driver": "En attente chauffeur", "disputed": "Contestée",
    "to_pay": "À payer", "paid": "Payée", "recharged": "Refacturée",
    "closed": "Clôturée", "cancelled": "Annulée",
}
TYPE_LABEL = {
    "speeding": "Excès de vitesse", "parking": "Stationnement",
    "red_light": "Feu rouge", "toll": "Péage",
    "forbidden_zone": "Zone interdite", "phone": "Téléphone",
    "seatbelt": "Ceinture", "other": "Autre",
}

# Columns reused by the three exporters (key, label)
COLUMNS = [
    ("dossier_number", "Dossier"),
    ("ref_fine", "Référence"),
    ("authority", "Autorité"),
    ("canton", "Canton"),
    ("city", "Commune"),
    ("infraction_at", "Date infraction"),
    ("location", "Lieu"),
    ("vehicle_plate", "Plaque"),
    ("driver_name", "Chauffeur"),
    ("driver_confidence", "Confiance %"),
    ("infraction_type", "Type"),
    ("amount", "Montant"),
    ("admin_fees", "Frais admin."),
    ("total_amount", "Total"),
    ("currency", "Devise"),
    ("due_date", "Échéance"),
    ("paid_at", "Payée le"),
    ("status", "Statut"),
    ("priority", "Priorité"),
]


def _fmt_dt(v: Optional[str]) -> str:
    if not v:
        return ""
    try:
        d = datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        return d.strftime("%d.%m.%Y %H:%M")
    except (ValueError, TypeError):
        return str(v)[:16]


def _fmt_date(v: Optional[str]) -> str:
    if not v:
        return ""
    try:
        d = datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        return d.strftime("%d.%m.%Y")
    except (ValueError, TypeError):
        return str(v)[:10]


def _cell(row: dict, key: str) -> str:
    v = row.get(key)
    if v is None or v == "":
        return ""
    if key == "infraction_at":
        return _fmt_dt(v)
    if key in ("due_date", "paid_at", "received_at"):
        return _fmt_date(v)
    if key == "status":
        return STATUS_LABEL.get(v, str(v))
    if key == "infraction_type":
        return TYPE_LABEL.get(v, str(v))
    if key in ("amount", "admin_fees", "total_amount"):
        try:
            return f"{float(v):.2f}"
        except (ValueError, TypeError):
            return str(v)
    return str(v)


def export_csv(rows: list[dict]) -> bytes:
    """RFC 4180 CSV with UTF-8 BOM (Excel-friendly)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([label for _, label in COLUMNS])
    for r in rows:
        writer.writerow([_cell(r, k) for k, _ in COLUMNS])
    # UTF-8 BOM so Excel opens accents correctly
    return "\ufeff".encode("utf-8") + buf.getvalue().encode("utf-8")


def export_excel(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Amendes"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")

    for ci, (_, label) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for ri, r in enumerate(rows, start=2):
        for ci, (k, _) in enumerate(COLUMNS, start=1):
            ws.cell(row=ri, column=ci, value=_cell(r, k))

    # Auto-fit-ish column widths (cap at 36)
    for ci, (k, label) in enumerate(COLUMNS, start=1):
        max_len = len(label)
        for r in rows:
            v = _cell(r, k)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[get_column_letter(ci)].width = min(36, max_len + 2)

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_pdf(rows: list[dict], totals: Optional[dict] = None) -> bytes:
    """Landscape A4 with a styled summary header + table."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        topMargin=14 * mm, bottomMargin=14 * mm,
        leftMargin=10 * mm, rightMargin=10 * mm,
        title="Logitrak — Gestion des amendes",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title", parent=styles["Title"], fontSize=16,
        textColor=colors.HexColor("#0f172a"), spaceAfter=6,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"], fontSize=9,
        textColor=colors.HexColor("#64748b"),
    )

    story = [
        Paragraph("Logitrak — Gestion des amendes", title_style),
        Paragraph(
            f"Export généré le {datetime.now(timezone.utc).strftime('%d.%m.%Y %H:%M')} UTC "
            f"&nbsp;·&nbsp; {len(rows)} amende{'s' if len(rows) > 1 else ''}",
            sub_style,
        ),
    ]
    if totals:
        story.append(Spacer(1, 4))
        story.append(Paragraph(
            f"<b>Montant total :</b> {totals.get('total_amount', 0):.2f} CHF "
            f"&nbsp;·&nbsp; <b>Payé :</b> {totals.get('paid_amount', 0):.2f} CHF "
            f"&nbsp;·&nbsp; <b>En attente :</b> {totals.get('open_amount', 0):.2f} CHF",
            sub_style,
        ))
    story.append(Spacer(1, 8))

    # PDF columns: pick the most important ones; full set fits on A4 landscape
    pdf_cols = [
        ("dossier_number", "Dossier"),
        ("infraction_at", "Date"),
        ("vehicle_plate", "Plaque"),
        ("driver_name", "Chauffeur"),
        ("infraction_type", "Type"),
        ("amount", "Montant"),
        ("total_amount", "Total"),
        ("due_date", "Échéance"),
        ("status", "Statut"),
    ]
    data = [[label for _, label in pdf_cols]]
    for r in rows:
        data.append([_cell(r, k) for k, _ in pdf_cols])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2196F3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("ALIGN", (5, 1), (6, -1), "RIGHT"),   # amounts
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#f8fafc")]),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#1976D2")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    doc.build(story)
    return buf.getvalue()
