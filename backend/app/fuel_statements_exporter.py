"""Exports Décomptes carburant — PDF (officiel), Excel (multi-onglets), CSV (comptable).

Toujours générés côté serveur, à partir du snapshot de lignes de la version courante.
Un décompte non clôturé porte la mention PROVISOIRE — ÉLÉMENTS À CONTRÔLER.
"""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.fuel_statements import STATUS_LABEL

CLASS_LABEL = {"professional": "Professionnel", "personal": "Privé", "unclassified": "Non classé"}
MATCH_LABEL = {"auto_matched": "Rapprochée", "matched_review": "Contrôle recommandé",
               "unmatched": "À rapprocher", "manual": "Attribuée manuellement"}

PROVISIONAL = "PROVISOIRE — ÉLÉMENTS À CONTRÔLER"

METHOD_TEXT = ("Méthode : conversion en CHF au taux de référence BCE de la date comptable "
               "(sinon date de transaction) dans le fuseau Europe/Zurich ; week-ends et jours fériés : "
               "dernier taux antérieur disponible. Le montant et la devise d'origine sont conservés. "
               "Répartition des coûts selon la classification des trajets rattachés (mode A). "
               "Numéros de cartes masqués — seuls les 4 derniers chiffres apparaissent.")


def _fmt(v, dec=2):
    return f"{v:.{dec}f}" if isinstance(v, (int, float)) else (v or "—")


def _version_note(stmt: dict) -> str | None:
    if stmt.get("version", 1) <= 1 or not stmt.get("versions"):
        return None
    prev = stmt["versions"][-1]
    diff = round((stmt.get("totals", {}).get("amount_chf_total") or 0)
                 - (prev.get("totals", {}).get("amount_chf_total") or 0), 2)
    return (f"Version corrigée V{stmt['version']} — remplace la V{prev['version']} "
            f"clôturée le {(prev.get('closed_at') or '')[:10]}. "
            f"Motif : {prev.get('replace_reason') or '—'}. "
            f"Écart total : {'+' if diff >= 0 else ''}{diff:.2f} CHF.")


def build_pdf(stmt: dict, lines: list[dict], tenant_name: str = "") -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=15, spaceAfter=2)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=11, spaceBefore=8, spaceAfter=3)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=7.5, textColor=colors.HexColor("#475569"))
    warn = ParagraphStyle("warn", parent=styles["Normal"], fontSize=12, textColor=colors.HexColor("#b91c1c"),
                          spaceBefore=4, spaceAfter=4)
    t = stmt.get("totals") or {}
    story = []
    title = f"Décompte carburant {stmt['number']} — V{stmt.get('version', 1)}"
    story.append(Paragraph(title, h1))
    meta = (f"{tenant_name + ' — ' if tenant_name else ''}Période : {stmt['date_from']} → {stmt['date_to']} · "
            f"Type : {'Correctif' if stmt.get('type') == 'corrective' else 'Régulier'} · "
            f"Statut : {STATUS_LABEL.get(stmt['status'], stmt['status'])} · "
            f"Généré le {(stmt.get('refreshed_at') or stmt.get('created_at') or '')[:16].replace('T', ' ')}"
            + (f" · Clôturé le {(stmt.get('closed_at') or '')[:16].replace('T', ' ')} par {stmt.get('closed_by')}"
               if stmt.get("status") == "closed" else ""))
    story.append(Paragraph(meta, small))
    if stmt.get("status") != "closed":
        story.append(Paragraph(f"<b>{PROVISIONAL}</b>", warn))
    vn = _version_note(stmt)
    if vn:
        story.append(Paragraph(vn, warn))
    if (stmt.get("close_exception") or {}).get("applied"):
        exc = stmt["close_exception"]
        story.append(Paragraph(
            f"Exception de clôture appliquée : {exc.get('excluded_count')} transaction(s) reportée(s). "
            f"Motif : {exc.get('reason')}", warn))

    # Synthèse
    story.append(Paragraph("Synthèse financière (CHF)", h2))
    b = t.get("blockers") or {}
    synth = [["Coût total CHF", "Transactions", "Litres", "kWh", "Professionnel", "Privé", "Non classé", "Éléments à contrôler"],
             [_fmt(t.get("amount_chf_total")), t.get("tx_count", 0), _fmt(t.get("liters")), _fmt(t.get("kwh")),
              _fmt(t.get("pro_chf")), _fmt(t.get("perso_chf")), _fmt(t.get("unclassified_chf")),
              b.get("total_count", 0)]]
    st = Table(synth, repeatRows=1)
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("ALIGN", (0, 1), (-1, -1), "CENTER"),
    ]))
    story.append(st)

    def _section_table(title_txt, headers, rows):
        story.append(Paragraph(title_txt, h2))
        if not rows:
            story.append(Paragraph("Aucune donnée", small))
            return
        tab = Table([headers] + rows, repeatRows=1)
        tab.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(tab)

    _section_table("Détail par véhicule",
                   ["Véhicule", "Transactions", "Litres", "kWh", "CHF", "Pro CHF", "Privé CHF"],
                   [[r["label"], r["tx_count"], _fmt(r["liters"]), _fmt(r["kwh"]),
                     _fmt(r["amount_chf"]), _fmt(r["pro_chf"]), _fmt(r["perso_chf"])]
                    for r in (t.get("by_vehicle") or [])])
    _section_table("Détail par chauffeur",
                   ["Chauffeur", "Transactions", "Litres", "kWh", "CHF", "Pro CHF", "Privé CHF"],
                   [[r["label"], r["tx_count"], _fmt(r["liters"]), _fmt(r["kwh"]),
                     _fmt(r["amount_chf"]), _fmt(r["pro_chf"]), _fmt(r["perso_chf"])]
                    for r in (t.get("by_driver") or [])])

    story.append(PageBreak())
    period = [l for l in lines if l["section"] == "period"]
    carried = [l for l in lines if l["section"] == "carried_over"]

    def _tx_rows(ls):
        rows = []
        for l in ls[:3000]:
            rows.append([
                l["basis_date"] + ("*" if l["basis"] == "accounting" else ""),
                f"•••• {l['card_last4']}" if l.get("card_last4") else "—",
                l.get("station_name") or "—",
                l.get("vehicle_plate") or "Non attribué",
                l.get("driver_name") or "—",
                f"{_fmt(l.get('quantity'))} {l.get('unit') or ''}" if l.get("quantity") else "—",
                f"{_fmt(l.get('amount_total'))} {l.get('currency')}",
                _fmt(l.get("amount_chf")) if l.get("amount_chf") is not None else "En attente",
                CLASS_LABEL.get(l.get("classification"), "—"),
                MATCH_LABEL.get(l.get("match_status"), l.get("match_status") or "—"),
            ])
        return rows

    tx_headers = ["Date*", "Carte", "Station", "Véhicule", "Chauffeur", "Quantité",
                  "Montant orig.", "CHF", "Classification", "Statut"]
    _section_table(f"Transactions de la période ({len(period)})", tx_headers, _tx_rows(period))
    if carried:
        _section_table(f"Transactions reportées / tardives incluses ({len(carried)})",
                       tx_headers, _tx_rows(carried))
    story.append(Spacer(1, 6))
    story.append(Paragraph("* Date comptable fournisseur lorsque disponible, sinon date de transaction.", small))
    story.append(Paragraph(METHOD_TEXT, small))
    doc.build(story)
    return buf.getvalue()


def build_excel(stmt: dict, lines: list[dict]) -> bytes:
    wb = Workbook()
    head_fill = PatternFill("solid", start_color="1E3A8A")
    head_font = Font(color="FFFFFF", bold=True, size=10)

    def _sheet(ws, headers, rows):
        ws.append(headers)
        for c in ws[1]:
            c.fill, c.font = head_fill, head_font
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, min(34, len(str(h)) + 6))

    t = stmt.get("totals") or {}
    b = t.get("blockers") or {}
    ws = wb.active
    ws.title = "Synthèse"
    rows = [
        ["Décompte", stmt["number"]], ["Version", f"V{stmt.get('version', 1)}"],
        ["Période", f"{stmt['date_from']} → {stmt['date_to']}"],
        ["Type", "Correctif" if stmt.get("type") == "corrective" else "Régulier"],
        ["Statut", STATUS_LABEL.get(stmt["status"], stmt["status"])
            + ("" if stmt["status"] == "closed" else f" — {PROVISIONAL}")],
        ["Clôturé le", (stmt.get("closed_at") or "—")[:19]],
        ["Coût total CHF", t.get("amount_chf_total")],
        ["Transactions", t.get("tx_count")],
        ["Litres", t.get("liters")], ["kWh", t.get("kwh")],
        ["Professionnel CHF", t.get("pro_chf")], ["Privé CHF", t.get("perso_chf")],
        ["Non classé CHF", t.get("unclassified_chf")],
        ["Non rapprochées", b.get("unmatched", {}).get("count", 0)],
        ["Conversions en attente", b.get("fx_pending", {}).get("count", 0)],
        ["Méthode", METHOD_TEXT],
    ]
    _sheet(ws, ["Champ", "Valeur"], rows)

    tx_headers = ["ID transaction", "Date utilisée", "Base", "Date transaction", "Date comptable",
                  "Fournisseur", "Carte", "Station", "Pays", "Produit", "Quantité", "Unité",
                  "Montant origine", "Devise", "TVA", "Taux appliqué", "Date du taux", "Source taux",
                  "Montant CHF", "Véhicule", "Chauffeur", "Trajet", "Classification",
                  "Statut rapprochement", "Section"]

    def _tx_row(l):
        return [l["transaction_id"], l["basis_date"], "Comptable" if l["basis"] == "accounting" else "Transaction",
                l.get("tx_datetime"), l.get("accounting_date"), l.get("provider"),
                f"•••• {l['card_last4']}" if l.get("card_last4") else "",
                l.get("station_name"), l.get("country"), l.get("product_type"),
                l.get("quantity"), l.get("unit"), l.get("amount_total"), l.get("currency"),
                l.get("vat_amount"), l.get("fx_rate"), l.get("fx_rate_date"), l.get("fx_source"),
                l.get("amount_chf"), l.get("vehicle_plate") or "Non attribué", l.get("driver_name"),
                l.get("trip_id"), CLASS_LABEL.get(l.get("classification"), ""),
                MATCH_LABEL.get(l.get("match_status"), l.get("match_status")),
                "Période" if l["section"] == "period" else "Reportée/tardive"]

    _sheet(wb.create_sheet("Transactions"), tx_headers, [_tx_row(l) for l in lines])
    agg_headers = ["Libellé", "Transactions", "Litres", "kWh", "CHF", "Pro CHF", "Privé CHF"]
    _sheet(wb.create_sheet("Par véhicule"), agg_headers,
           [[r["label"], r["tx_count"], r["liters"], r["kwh"], r["amount_chf"], r["pro_chf"], r["perso_chf"]]
            for r in (t.get("by_vehicle") or [])])
    _sheet(wb.create_sheet("Par chauffeur"), agg_headers,
           [[r["label"], r["tx_count"], r["liters"], r["kwh"], r["amount_chf"], r["pro_chf"], r["perso_chf"]]
            for r in (t.get("by_driver") or [])])
    _sheet(wb.create_sheet("Privé-Professionnel"),
           ["Classification", "Montant CHF"],
           [["Professionnel", t.get("pro_chf")], ["Privé", t.get("perso_chf")],
            ["Non classé", t.get("unclassified_chf")]])
    _sheet(wb.create_sheet("Reportées"), tx_headers,
           [_tx_row(l) for l in lines if l["section"] == "carried_over"])
    fx_used = sorted({(l.get("currency"), l.get("fx_rate"), l.get("fx_rate_date"), l.get("fx_source"))
                      for l in lines if l.get("fx_rate") is not None and l.get("currency") != "CHF"})
    _sheet(wb.create_sheet("Taux de change"),
           ["Devise", "Taux (CHF pour 1 unité)", "Date du taux", "Source"],
           [[c, r, d, ("BCE" if s == "ecb" else s)] for c, r, d, s in fx_used])
    _sheet(wb.create_sheet("Historique"),
           ["Version", "Clôturé le", "Par", "Total CHF", "Statut", "Motif remplacement"],
           [[f"V{v['version']}", (v.get("closed_at") or "")[:19], v.get("closed_by"),
             (v.get("totals") or {}).get("amount_chf_total"), "Annulée et remplacée",
             v.get("replace_reason")]
            for v in (stmt.get("versions") or [])]
           + [[f"V{stmt.get('version', 1)}", (stmt.get("closed_at") or "")[:19], stmt.get("closed_by"),
               t.get("amount_chf_total"), STATUS_LABEL.get(stmt["status"]), ""]])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_csv(stmt: dict, lines: list[dict]) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["transaction_id", "statement_number", "statement_version", "basis_date", "basis",
                "tx_datetime", "accounting_date", "provider", "card_masked", "station", "country",
                "product_type", "quantity", "unit", "amount_original", "currency", "vat_amount",
                "fx_rate", "fx_rate_date", "fx_source", "amount_chf", "vehicle_plate", "driver_name",
                "trip_id", "classification", "match_status", "section"])
    for l in lines:
        w.writerow([l["transaction_id"], stmt["number"], f"V{stmt.get('version', 1)}",
                    l["basis_date"], l["basis"], l.get("tx_datetime"), l.get("accounting_date"),
                    l.get("provider"), f"****{l['card_last4']}" if l.get("card_last4") else "",
                    l.get("station_name"), l.get("country"), l.get("product_type"),
                    l.get("quantity"), l.get("unit"), l.get("amount_total"), l.get("currency"),
                    l.get("vat_amount"), l.get("fx_rate"), l.get("fx_rate_date"), l.get("fx_source"),
                    l.get("amount_chf"), l.get("vehicle_plate"), l.get("driver_name"),
                    l.get("trip_id"), l.get("classification"), l.get("match_status"), l["section"]])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")
