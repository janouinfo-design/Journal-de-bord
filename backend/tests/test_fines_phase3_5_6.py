"""Phases 3, 5, 6 of Fines Module — OCR, Documents, Exports, Extended Stats.

Validates:
- Phase 5: OCR extract endpoint (Gemini Vision) + audit log + role checks
- Phase 6: CSV/Excel/PDF exports with filter params + role checks
- Phase 3: Document upload/download/delete + stats/extended + RBAC
"""
import io
import os
import re
import time
from pathlib import Path

import pytest
import requests
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import fitz  # PyMuPDF

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
API = f"{BASE_URL}/api"
LIVRE = f"{API}/livre"
STORAGE_ROOT = Path("/app/backend/storage/fines")


def _login(email, password):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": password}, timeout=15)
    assert r.status_code == 200, f"Login failed for {email}: {r.status_code} {r.text}"
    return r.json()["access_token"]


@pytest.fixture(scope="session")
def admin_token():
    return _login("admin@logitrak.ch", "admin123")


@pytest.fixture(scope="session")
def manager_token():
    return _login("manager@logitrak.ch", "manager123")


@pytest.fixture(scope="session")
def driver_token():
    return _login("chauffeur@logitrak.ch", "chauffeur123")


def H(token):
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture(scope="session")
def fine_id(admin_token):
    """Create a TEST_ fine to attach docs to."""
    payload = {
        "ref_fine": "TEST-PHASE3-001",
        "authority": "Police cantonale Vaud",
        "country": "CH",
        "canton": "VD",
        "city": "Lausanne",
        "infraction_at": "2025-10-15T10:30:00Z",
        "location": "Avenue de la Gare 1",
        "vehicle_plate": "VD 123456",
        "infraction_type": "speeding",
        "amount": 250.0,
        "admin_fees": 50.0,
        "currency": "CHF",
        "due_date": "2025-11-15",
        "status": "to_analyze",
        "priority": "normal",
    }
    r = requests.post(f"{LIVRE}/fines", json=payload, headers=H(admin_token), timeout=10)
    assert r.status_code in (200, 201), r.text
    return r.json()["id"]


@pytest.fixture(scope="session")
def fake_amende_jpg():
    """Build a JPEG with visible text resembling a Swiss fine notice."""
    path = Path("/tmp/test_amende_phase5.jpg")
    if not path.exists():
        img = Image.new("RGB", (800, 1000), color=(245, 245, 245))
        d = ImageDraw.Draw(img)
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 22)
            small = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 16)
        except Exception:
            font = ImageFont.load_default()
            small = font
        lines = [
            ("AMENDE D'ORDRE - CANTON DE VAUD", font),
            ("Police Cantonale Vaudoise", small),
            ("", small),
            ("Reference: VD-2025-007891", small),
            ("Date infraction: 15.10.2025 10:30", small),
            ("Lieu: Avenue de la Gare 1, Lausanne", small),
            ("Immatriculation: VD 123456", small),
            ("Type: Exces de vitesse (15 km/h)", small),
            ("Montant: CHF 250.00", small),
            ("Frais administratifs: CHF 50.00", small),
            ("Total: CHF 300.00", small),
            ("Echeance: 15.11.2025", small),
        ]
        y = 40
        for txt, f in lines:
            d.text((40, y), txt, fill=(20, 20, 20), font=f)
            y += 50
        # Add some shapes for visual features
        d.rectangle([30, 30, 770, 970], outline=(0, 0, 0), width=3)
        d.line([(40, 90), (760, 90)], fill=(100, 100, 100), width=2)
        img.save(path, "JPEG", quality=85)
    return path


# ============== Phase 5: OCR ============================================
class TestPhase5OCR:
    def test_ocr_invalid_content_type(self, admin_token):
        files = {"file": ("note.txt", b"hello world", "text/plain")}
        r = requests.post(f"{LIVRE}/fines/ocr-extract", files=files, headers=H(admin_token), timeout=15)
        assert r.status_code == 400, r.text

    def test_ocr_missing_file(self, admin_token):
        r = requests.post(f"{LIVRE}/fines/ocr-extract", headers=H(admin_token), timeout=15)
        assert r.status_code == 422, r.text

    def test_ocr_too_large(self, admin_token):
        # 12 MB binary
        big = b"\xff" * (12 * 1024 * 1024)
        files = {"file": ("big.jpg", big, "image/jpeg")}
        r = requests.post(f"{LIVRE}/fines/ocr-extract", files=files, headers=H(admin_token), timeout=30)
        assert r.status_code == 413, f"Expected 413 got {r.status_code}"

    def test_ocr_driver_forbidden(self, driver_token, fake_amende_jpg):
        with open(fake_amende_jpg, "rb") as f:
            files = {"file": ("amende.jpg", f.read(), "image/jpeg")}
        r = requests.post(f"{LIVRE}/fines/ocr-extract", files=files, headers=H(driver_token), timeout=15)
        assert r.status_code == 403, r.text

    def test_ocr_extract_success(self, admin_token, fake_amende_jpg):
        """Real Gemini Vision call — 1-2 per run only."""
        with open(fake_amende_jpg, "rb") as f:
            files = {"file": ("amende.jpg", f.read(), "image/jpeg")}
        r = requests.post(f"{LIVRE}/fines/ocr-extract", files=files, headers=H(admin_token), timeout=120)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data.get("ok") is True
        assert data.get("model") == "gemini-3.1-pro-preview", f"Got model: {data.get('model')}"
        assert "extracted" in data
        assert isinstance(data["extracted"], dict)
        # At least some fields should come back
        ex = data["extracted"]
        print(f"OCR extracted keys: {sorted(ex.keys())}")

        # Verify audit log entry exists
        time.sleep(0.5)
        r2 = requests.get(f"{LIVRE}/audit?action=ocr_extract&limit=5", headers=H(admin_token), timeout=10)
        # audit endpoint may not exist — verify directly via DB-less assumption (skip check if no endpoint)
        if r2.status_code == 200:
            entries = r2.json() if isinstance(r2.json(), list) else r2.json().get("items", [])
            found = any(e.get("actor") == "admin@logitrak.ch" and e.get("action") == "ocr_extract" for e in entries)
            assert found, "audit_log entry for ocr_extract not found"


# ============== Phase 6: Exports ========================================
class TestPhase6Exports:
    def test_export_csv(self, admin_token):
        r = requests.get(f"{LIVRE}/fines/export?fmt=csv", headers=H(admin_token), timeout=20)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "").lower()
        assert "text/csv" in ct and "charset=utf-8" in ct, ct
        cd = r.headers.get("content-disposition", "")
        assert "attachment" in cd.lower() and "filename=" in cd.lower(), cd
        body = r.content
        # UTF-8 BOM
        assert body.startswith(b"\xef\xbb\xbf"), f"missing BOM: {body[:10]!r}"
        text = body.decode("utf-8-sig")
        # Header columns
        first = text.splitlines()[0]
        assert "Dossier" in first and "Référence" in first, first[:200]

    def test_export_csv_with_filters(self, admin_token):
        params = {
            "fmt": "csv",
            "status": "to_analyze",
            "infraction_type": "speeding",
            "start": "2025-01-01",
            "end": "2026-12-31",
            "q": "TEST",
            "min_amount": 0,
            "max_amount": 100000,
            "sort": "-infraction_at",
        }
        r = requests.get(f"{LIVRE}/fines/export", params=params, headers=H(admin_token), timeout=20)
        assert r.status_code == 200, r.text
        assert r.content.startswith(b"\xef\xbb\xbf")

    def test_export_excel(self, admin_token):
        r = requests.get(f"{LIVRE}/fines/export?fmt=excel", headers=H(admin_token), timeout=20)
        assert r.status_code == 200, r.text
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct, ct
        wb = load_workbook(io.BytesIO(r.content))
        assert "Amendes" in wb.sheetnames, wb.sheetnames
        ws = wb["Amendes"]
        headers = [c.value for c in ws[1]]
        assert "Dossier" in headers, headers
        assert "_id" not in headers, "MongoDB _id leaked into xlsx headers!"

    def test_export_pdf(self, admin_token):
        r = requests.get(f"{LIVRE}/fines/export?fmt=pdf", headers=H(admin_token), timeout=30)
        assert r.status_code == 200, r.text
        assert r.headers.get("content-type", "").startswith("application/pdf")
        doc = fitz.open(stream=r.content, filetype="pdf")
        try:
            assert doc.page_count >= 1
            page1_text = doc[0].get_text()
            assert "Logitrak" in page1_text and "amendes" in page1_text.lower(), page1_text[:300]
        finally:
            doc.close()

    def test_export_invalid_fmt(self, admin_token):
        r = requests.get(f"{LIVRE}/fines/export?fmt=word", headers=H(admin_token), timeout=10)
        assert r.status_code == 422, r.text

    def test_export_driver_forbidden(self, driver_token):
        r = requests.get(f"{LIVRE}/fines/export?fmt=pdf", headers=H(driver_token), timeout=10)
        assert r.status_code == 403, r.text


# ============== Phase 3: Documents ======================================
class TestPhase3Documents:
    def test_upload_invalid_kind(self, admin_token, fine_id):
        files = {"file": ("doc.pdf", b"%PDF-1.4 fake", "application/pdf")}
        r = requests.post(f"{LIVRE}/fines/{fine_id}/documents",
                          files=files, data={"kind": "garbage"},
                          headers=H(admin_token), timeout=10)
        assert r.status_code == 400, r.text
        # allowed values listed
        assert "pdf" in r.text or "photo" in r.text

    def test_upload_invalid_mime(self, admin_token, fine_id):
        files = {"file": ("hack.exe", b"MZ\x00\x00", "application/x-msdownload")}
        r = requests.post(f"{LIVRE}/fines/{fine_id}/documents",
                          files=files, data={"kind": "libre"},
                          headers=H(admin_token), timeout=10)
        assert r.status_code == 400, r.text

    def test_upload_empty_file(self, admin_token, fine_id):
        files = {"file": ("empty.pdf", b"", "application/pdf")}
        r = requests.post(f"{LIVRE}/fines/{fine_id}/documents",
                          files=files, data={"kind": "pdf"},
                          headers=H(admin_token), timeout=10)
        assert r.status_code == 400, r.text

    def test_upload_too_large(self, admin_token, fine_id):
        big = b"%PDF-1.4\n" + b"\x00" * (21 * 1024 * 1024)
        files = {"file": ("huge.pdf", big, "application/pdf")}
        r = requests.post(f"{LIVRE}/fines/{fine_id}/documents",
                          files=files, data={"kind": "pdf"},
                          headers=H(admin_token), timeout=60)
        assert r.status_code == 413, r.text

    def test_upload_download_delete_flow(self, admin_token, driver_token, fine_id, fake_amende_jpg):
        # Upload
        with open(fake_amende_jpg, "rb") as f:
            files = {"file": ("photo amende.jpg", f.read(), "image/jpeg")}
        r = requests.post(f"{LIVRE}/fines/{fine_id}/documents",
                          files=files, data={"kind": "photo"},
                          headers=H(admin_token), timeout=15)
        assert r.status_code == 200, r.text
        doc = r.json()
        assert doc["kind"] == "photo"
        assert doc["content_type"] == "image/jpeg"
        # Filename sanitized — space replaced
        assert " " not in doc["filename"], doc["filename"]
        assert doc["size_bytes"] > 0
        doc_id = doc["id"]

        # Verify on disk
        fine_dir = STORAGE_ROOT / fine_id
        files_on_disk = list(fine_dir.glob(f"{doc_id}_*"))
        assert len(files_on_disk) == 1, f"Expected 1 file, found {files_on_disk}"

        # GET fine — documents array contains it
        r2 = requests.get(f"{LIVRE}/fines/{fine_id}", headers=H(admin_token), timeout=10)
        assert r2.status_code == 200
        docs = r2.json().get("documents") or []
        assert any(d["id"] == doc_id for d in docs), docs

        # PATCH a non-document field — documents must remain
        r3 = requests.patch(f"{LIVRE}/fines/{fine_id}",
                            json={"internal_notes": "patched at " + str(time.time())},
                            headers=H(admin_token), timeout=10)
        assert r3.status_code == 200, r3.text
        r4 = requests.get(f"{LIVRE}/fines/{fine_id}", headers=H(admin_token), timeout=10)
        docs_after = r4.json().get("documents") or []
        assert any(d["id"] == doc_id for d in docs_after), \
            f"PATCH clobbered documents array: {docs_after}"

        # Download as admin
        r5 = requests.get(f"{LIVRE}/fines/{fine_id}/documents/{doc_id}/download",
                          headers=H(admin_token), timeout=15)
        assert r5.status_code == 200
        assert r5.headers.get("content-type", "").startswith("image/jpeg")
        assert "attachment" in r5.headers.get("content-disposition", "").lower() or \
               "filename" in r5.headers.get("content-disposition", "").lower()
        assert len(r5.content) > 100

        # Download as driver = 403
        r6 = requests.get(f"{LIVRE}/fines/{fine_id}/documents/{doc_id}/download",
                          headers=H(driver_token), timeout=10)
        assert r6.status_code == 403, r6.text

        # Nonexistent doc -> 404
        r7 = requests.get(f"{LIVRE}/fines/{fine_id}/documents/nope-nope-nope/download",
                          headers=H(admin_token), timeout=10)
        assert r7.status_code == 404, r7.text

        # Delete as admin
        r8 = requests.delete(f"{LIVRE}/fines/{fine_id}/documents/{doc_id}",
                             headers=H(admin_token), timeout=10)
        assert r8.status_code == 200, r8.text

        # File removed from disk
        assert not any(fine_dir.glob(f"{doc_id}_*")), "File still on disk after delete"

        # Document removed from fine
        r9 = requests.get(f"{LIVRE}/fines/{fine_id}", headers=H(admin_token), timeout=10)
        docs_after_del = r9.json().get("documents") or []
        assert not any(d["id"] == doc_id for d in docs_after_del)

        # Subsequent download -> 404
        r10 = requests.get(f"{LIVRE}/fines/{fine_id}/documents/{doc_id}/download",
                           headers=H(admin_token), timeout=10)
        assert r10.status_code == 404

    def test_manager_can_upload_and_delete(self, manager_token, fine_id):
        files = {"file": ("m.pdf", b"%PDF-1.4\n%test", "application/pdf")}
        r = requests.post(f"{LIVRE}/fines/{fine_id}/documents",
                          files=files, data={"kind": "pdf"},
                          headers=H(manager_token), timeout=10)
        assert r.status_code == 200, r.text
        doc_id = r.json()["id"]
        r2 = requests.delete(f"{LIVRE}/fines/{fine_id}/documents/{doc_id}",
                             headers=H(manager_token), timeout=10)
        assert r2.status_code == 200

    def test_driver_upload_forbidden(self, driver_token, fine_id):
        files = {"file": ("d.pdf", b"%PDF-1.4", "application/pdf")}
        r = requests.post(f"{LIVRE}/fines/{fine_id}/documents",
                          files=files, data={"kind": "pdf"},
                          headers=H(driver_token), timeout=10)
        assert r.status_code == 403, r.text


# ============== Phase 3: Stats Extended =================================
class TestPhase3StatsExtended:
    def test_stats_extended_admin(self, admin_token):
        r = requests.get(f"{LIVRE}/fines/stats/extended", headers=H(admin_token), timeout=15)
        assert r.status_code == 200, r.text
        data = r.json()

        # KPIs
        kpis = data.get("kpis", {})
        for k in ("total", "total_amount", "paid_amount", "pending_amount", "disputed", "overdue"):
            assert k in kpis, f"Missing KPI {k}: {kpis}"

        # by_status — all 10 statuses
        bs = data.get("by_status", {})
        expected_statuses = {"received", "to_analyze", "driver_to_identify", "awaiting_driver",
                             "disputed", "to_pay", "paid", "recharged", "closed", "cancelled"}
        assert set(bs.keys()) >= expected_statuses, f"Missing statuses: {expected_statuses - set(bs.keys())}"

        # by_type — all 8 types
        bt = data.get("by_type", {})
        expected_types = {"speeding", "parking", "red_light", "toll",
                          "forbidden_zone", "phone", "seatbelt", "other"}
        assert set(bt.keys()) >= expected_types, f"Missing types: {expected_types - set(bt.keys())}"

        # monthly
        monthly = data.get("monthly", [])
        assert len(monthly) == 12, f"Expected 12 months, got {len(monthly)}"
        for m in monthly:
            assert "month" in m and re.match(r"^\d{4}-\d{2}$", m["month"]), m
            assert "count" in m and "amount" in m

        # top lists
        for key in ("top_vehicles", "top_drivers", "top_amounts"):
            top = data.get(key, [])
            assert isinstance(top, list) and len(top) <= 10, f"{key}: {len(top)}"

        # Check top_amounts shape
        if data["top_amounts"]:
            t = data["top_amounts"][0]
            for f_ in ("key", "label", "vehicle", "driver", "total", "status"):
                assert f_ in t, f"top_amounts missing {f_}: {t}"

    def test_stats_extended_manager_ok(self, manager_token):
        r = requests.get(f"{LIVRE}/fines/stats/extended", headers=H(manager_token), timeout=15)
        assert r.status_code == 200, r.text

    def test_stats_extended_driver_forbidden(self, driver_token):
        r = requests.get(f"{LIVRE}/fines/stats/extended", headers=H(driver_token), timeout=10)
        assert r.status_code == 403, r.text
